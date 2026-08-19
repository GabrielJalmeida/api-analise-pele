from __future__ import annotations

import csv
import io
import re
import unicodedata

from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from database import gerenciar_transacao
from models import ConfirmarImportacaoCatalogo, NovoProduto


TAMANHO_MAXIMO_ARQUIVO = 5 * 1024 * 1024
LIMITE_LINHAS = 1000


class ArquivoCatalogoInvalido(ValueError):
    pass


ALIASES_CAMPOS = {
    "nome": "nome",
    "produto": "nome",
    "nome produto": "nome",
    "marca": "marca",
    "descricao": "descricao_curta",
    "descricao curta": "descricao_curta",
    "imagem": "imagem_url",
    "imagem url": "imagem_url",
    "url imagem": "imagem_url",
    "conteudo": "conteudo",
    "volume": "conteudo",
    "quantidade embalagem": "conteudo",
    "ativos": "ativos_principais",
    "ativos principais": "ativos_principais",
    "ingredientes ativos": "ativos_principais",
    "preco": "preco",
    "valor": "preco",
    "estoque": "estoque",
    "quantidade": "estoque",
    "categoria": "categoria",
    "tipo pele": "tipo_pele",
    "tipo de pele": "tipo_pele",
    "pele sensivel": "pele_sensivel",
    "sensivel": "pele_sensivel",
    "indicado para espinha": "indicado_para_espinha",
    "espinha": "indicado_para_espinha",
    "acne": "indicado_para_espinha",
    "ativo": "ativo",
    "status": "ativo",
}


def _normalizar_texto(valor: object) -> str:
    texto = (
        ""
        if valor is None
        else str(valor)
    ).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )
    texto = re.sub(r"[_\-/]+", " ", texto)
    return " ".join(texto.split())


def _texto(valor: object) -> str:
    if valor is None:
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()


def _preco(valor: object) -> float | None:
    if valor is None or str(valor).strip() == "":
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    texto = re.sub(r"[^0-9,.-]", "", texto)

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return float(texto)
    except ValueError as erro:
        raise ValueError("Informe um preço válido") from erro


def _inteiro(valor: object, padrao: int = 0) -> int:
    if valor is None or str(valor).strip() == "":
        return padrao

    try:
        numero = float(str(valor).replace(",", "."))
    except ValueError as erro:
        raise ValueError("Informe um estoque inteiro válido") from erro

    if not numero.is_integer():
        raise ValueError("O estoque precisa ser um número inteiro")

    return int(numero)


def _booleano(valor: object, padrao: bool) -> bool:
    if valor is None or str(valor).strip() == "":
        return padrao

    texto = _normalizar_texto(valor)

    if texto in {"1", "sim", "s", "true", "verdadeiro", "ativo"}:
        return True

    if texto in {"0", "nao", "n", "false", "falso", "inativo"}:
        return False

    raise ValueError("Use sim ou não")


def _categoria(valor: object) -> str | None:
    texto = _normalizar_texto(valor)

    if not texto:
        return None

    aliases = {
        "limpeza": "limpeza",
        "limpador": "limpeza",
        "sabonete": "limpeza",
        "hidratante": "hidratante",
        "hidratacao": "hidratante",
        "serum": "serum",
        "tratamento": "serum",
        "protetor": "protetor_solar",
        "protetor solar": "protetor_solar",
        "protecao solar": "protetor_solar",
        "outro": "outros",
        "outros": "outros",
    }

    return aliases.get(texto, texto.replace(" ", "_"))


def _tipo_pele(valor: object) -> str | None:
    texto = _normalizar_texto(valor)

    if not texto:
        return None

    aliases = {
        "oleosa": "oleosa",
        "seca": "seca",
        "mista": "mista",
        "normal": "normal",
        "todos": "todos",
        "todas": "todos",
        "todos os tipos": "todos",
        "todas as peles": "todos",
    }

    return aliases.get(texto, texto)


def _linhas_csv(conteudo: bytes) -> Iterable[list[object]]:
    try:
        texto = conteudo.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            texto = conteudo.decode("latin-1")
        except UnicodeDecodeError as erro:
            raise ArquivoCatalogoInvalido(
                "Não foi possível ler a codificação do CSV"
            ) from erro

    amostra = texto[:4096]

    try:
        dialeto = csv.Sniffer().sniff(
            amostra,
            delimiters=",;\t",
        )
    except csv.Error:
        dialeto = csv.excel

    return list(csv.reader(io.StringIO(texto), dialect=dialeto))


def _linhas_xlsx(conteudo: bytes) -> Iterable[list[object]]:
    try:
        pasta = load_workbook(
            io.BytesIO(conteudo),
            read_only=True,
            data_only=True,
        )
    except Exception as erro:
        raise ArquivoCatalogoInvalido(
            "A planilha XLSX é inválida ou está corrompida"
        ) from erro

    try:
        planilha = pasta.active
        return [list(linha) for linha in planilha.iter_rows(values_only=True)]
    finally:
        pasta.close()


def _mapear_cabecalho(cabecalho: list[object]) -> dict[int, str]:
    mapeamento = {}

    for indice, valor in enumerate(cabecalho):
        campo = ALIASES_CAMPOS.get(
            _normalizar_texto(valor)
        )

        if campo:
            mapeamento[indice] = campo

    if "nome" not in mapeamento.values():
        raise ArquivoCatalogoInvalido(
            "A planilha precisa ter uma coluna 'nome' ou 'produto'"
        )

    return mapeamento


def _preparar_produto(dados: dict[str, object]) -> dict:
    return {
        "nome": _texto(dados.get("nome")),
        "marca": _texto(dados.get("marca")),
        "descricao_curta": _texto(dados.get("descricao_curta")),
        "imagem_url": _texto(dados.get("imagem_url")),
        "conteudo": _texto(dados.get("conteudo")),
        "ativos_principais": _texto(dados.get("ativos_principais")),
        "preco": _preco(dados.get("preco")),
        "estoque": _inteiro(dados.get("estoque")),
        "categoria": _categoria(dados.get("categoria")),
        "tipo_pele": _tipo_pele(dados.get("tipo_pele")),
        "pele_sensivel": _booleano(
            dados.get("pele_sensivel"),
            False,
        ),
        "indicado_para_espinha": _booleano(
            dados.get("indicado_para_espinha"),
            False,
        ),
        "ativo": _booleano(
            dados.get("ativo"),
            True,
        ),
    }


def validar_rascunhos(
    rascunhos: list[dict],
    *,
    primeira_linha: int = 1,
) -> dict:
    produtos = []
    erros = []

    for indice, rascunho in enumerate(rascunhos):
        linha = primeira_linha + indice

        try:
            preparado = _preparar_produto(rascunho)
            produto = NovoProduto.model_validate(preparado)
            produtos.append(produto)
        except (ValueError, TypeError) as erro:
            if isinstance(erro, ValidationError):
                for detalhe in erro.errors():
                    erros.append(
                        {
                            "linha": linha,
                            "campo": ".".join(
                                str(parte)
                                for parte in detalhe["loc"]
                            ),
                            "mensagem": detalhe["msg"],
                        }
                    )
            else:
                erros.append(
                    {
                        "linha": linha,
                        "campo": "linha",
                        "mensagem": str(erro),
                    }
                )

    return {
        "produtos": produtos,
        "erros": erros,
        "total_linhas": len(rascunhos),
    }


def processar_arquivo_catalogo(
    conteudo: bytes,
    nome_arquivo: str,
) -> dict:
    if not conteudo:
        raise ArquivoCatalogoInvalido("O arquivo está vazio")

    if len(conteudo) > TAMANHO_MAXIMO_ARQUIVO:
        raise ArquivoCatalogoInvalido(
            "O arquivo não pode ultrapassar 5 MB"
        )

    extensao = Path(nome_arquivo).suffix.lower()

    if extensao == ".csv":
        linhas = list(_linhas_csv(conteudo))
    elif extensao == ".xlsx":
        linhas = list(_linhas_xlsx(conteudo))
    else:
        raise ArquivoCatalogoInvalido(
            "Envie um arquivo CSV ou XLSX"
        )

    linhas = [
        linha
        for linha in linhas
        if any(_texto(valor) for valor in linha)
    ]

    if len(linhas) < 2:
        raise ArquivoCatalogoInvalido(
            "A planilha precisa de cabeçalho e pelo menos um produto"
        )

    if len(linhas) - 1 > LIMITE_LINHAS:
        raise ArquivoCatalogoInvalido(
            "A planilha não pode ultrapassar 1.000 produtos"
        )

    mapeamento = _mapear_cabecalho(linhas[0])
    rascunhos = []

    for linha in linhas[1:]:
        rascunhos.append(
            {
                campo: (
                    linha[indice]
                    if indice < len(linha)
                    else None
                )
                for indice, campo in mapeamento.items()
            }
        )

    return validar_rascunhos(
        rascunhos,
        primeira_linha=2,
    )


def confirmar_importacao(
    dados: ConfirmarImportacaoCatalogo,
) -> dict:
    criados = 0
    atualizados = 0
    ignorados = 0

    with gerenciar_transacao() as (_, cursor):
        for produto in dados.produtos:
            cursor.execute(
                """
                SELECT id FROM produtos
                WHERE nome = ? COLLATE NOCASE
                """,
                (produto.nome,),
            )
            existente = cursor.fetchone()
            valores = (
                produto.preco,
                produto.estoque,
                produto.tipo_pele,
                produto.pele_sensivel,
                produto.indicado_para_espinha,
                produto.ativo,
                produto.categoria,
                produto.marca,
                produto.descricao_curta,
                produto.imagem_url,
                produto.conteudo,
                produto.ativos_principais,
            )

            if existente and dados.duplicados == "ignorar":
                ignorados += 1
                continue

            if existente:
                cursor.execute(
                    """
                    UPDATE produtos SET
                        preco = ?, estoque = ?, tipo_pele = ?,
                        pele_sensivel = ?, indicado_para_espinha = ?,
                        ativo = ?, categoria = ?, marca = ?,
                        descricao_curta = ?, imagem_url = ?,
                        conteudo = ?, ativos_principais = ?
                    WHERE id = ?
                    """,
                    (*valores, existente["id"]),
                )
                atualizados += 1
                continue

            cursor.execute(
                """
                INSERT INTO produtos (
                    nome, preco, estoque, tipo_pele,
                    pele_sensivel, indicado_para_espinha,
                    ativo, categoria, marca, descricao_curta,
                    imagem_url, conteudo, ativos_principais
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (produto.nome, *valores),
            )
            criados += 1

    return {
        "status": "importacao_concluida",
        "criados": criados,
        "atualizados": atualizados,
        "ignorados": ignorados,
    }
